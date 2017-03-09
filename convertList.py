#coding=utf-8
import sys
reload(sys)
sys.setdefaultencoding('utf8')
import codecs
import types
import chardet
from openpyxl import load_workbook

from xml.dom.minidom import parse
import xml.dom.minidom


try: 
    import xml.etree.cElementTree as ET 
except ImportError: 
    import xml.etree.ElementTree as ET 




def read_excel():

	ipath = u'I:/proj_001_20170124_xct/Tool/xls_convert/xls/configxls/副本系统1.0/1.0副本剧情表.xlsx'
	# uipath = unicode(ipath , "utf8")
	# f = open(ipath)
	# print f
	wb = load_workbook(ipath)	 
	# # wb = load_workbook('I:/proj_001_20170124_xct/Tool/xls_convert/xls/abc.xlsx', "utf-8")
	# # wb = load_workbook('I:/proj_001_20170124_xct/Tool/xls_convert/xls/abc.xlsx')

	sheets = wb.get_sheet_names()   # 获取所有表格(worksheet)的名字
	sheet0 = sheets[0]  # 第一个表格的名称
	ws = wb.get_sheet_by_name(sheet0) # 获取特定的 worksheet

	# 获取表格所有行和列，两者都是可迭代的
	rows = ws.rows
	columns = ws.columns
	 
	# 行迭代
	content = []
	for row in rows:
	    line = [col.value for col in row]
	    # print line
	    content.append(line)

	f = codecs.open("./juqing.txt","w+","utf-8")
	f.write("{\n")
	writeIndex = 0
	for obj in content:
		# print type(obj)
		for subobj in obj:
			print subobj
			# print type(subobj)
			if type(subobj) is types.UnicodeType :
				# print "heheheh*_++++++++++++"
				f.write(subobj)


		pass
		# f.write()
	f.write("\n}\n")

	# # 通过坐标读取值
	# print ws.cell('B12').value    # B 表示列，12 表示行
	# print ws.cell(row=12, column=2).value





def get_charset(s):
	return chardet.detect(s)['encoding']


macroDict = {}
structDict = {}
sortkeyDict = {}

def parseResouceXml():
	print("parseResouceXml")

	fileName = "./xml/resource.xml"
	file_xml = open(fileName,"r").read()
	charset = get_charset(file_xml)
	print "old charset:"+charset
	file_xml = file_xml.decode(charset).encode("utf-8")
	# print get_charset(file_xml)
	try: 
		MetaRoot = ET.fromstring(file_xml)     #打开xml文档 
	except Exception, e: 
		print e
		print "Error:cannot parse file:." + fileName
		return


	for child in MetaRoot:
		# print child.tag


		if child.tag == "macro":
			Name =  (child.attrib)["name"]
			Value =  (child.attrib)["value"]
			macroDict[Name] = Value
		elif child.tag == "struct":
			structItemDict = []
			sortkey = (child.attrib)["sortkey"]
			sortkeyCh = ""
			sortkeyChType = ""
			for item in child:
				# print item.attrib["name"]
				# print item.attrib["type"]
				# print item.attrib["cname"]
				itemDict = {}
				cname = item.attrib["cname"] 
				itemDict["name"] = item.attrib["name"]
				itemDict["type"] = item.attrib["type"]
				itemDict["cname"] = item.attrib["cname"]
				if sortkey == itemDict["name"]:
					sortkeyCh = itemDict["cname"]
					sortkeyChType = itemDict["type"]

				if item.attrib.has_key("size"):
					# print "size----------------",item.attrib["size"], macroDict[item.attrib["size"]]
					itemDict["size"] = macroDict[item.attrib["size"]]
				# if item.attrib["cname"]
				# itemDict["size"] = 
				# structItemDict[cname] = itemDict
				structItemDict.append(itemDict)

			structDict[child.attrib["name"]] = structItemDict
			sortkeyDict[child.attrib["name"]] = {"type":sortkeyChType ,"cname":sortkeyCh}

	# print macroDict
	# print structDict
	print sortkeyDict


outputFolder = "./config/"
def xlsToXml(xlsxFile, meta):
	# print xlsxFile
	inputXlsxFile = ".\\xls\\configxls\\" + xlsxFile
	print inputXlsxFile
	dstXmlFile =  outputFolder+meta+".xml"
	print dstXmlFile
	f = codecs.open(dstXmlFile,"w+","utf-8")
	impl = xml.dom.minidom.getDOMImplementation()
	dom = impl.createDocument(None,  meta + '_Tab', None)

	root = dom.documentElement
	headE = dom.createElement('TResHeadAll')
	root.appendChild(headE)
	resheadE = dom.createElement('resHead')
	headE.appendChild(resheadE)

	resheadE.setAttribute("type","TResHead")

	# root = ET.Element("root")
	# root.tail = "\n"
	# ET.SubElement(root, "TResHeadAll")
	# tree = ET.ElementTree(root)
	# tree.write(dstXmlFile,encoding='utf-8')
	# pass
	wb = load_workbook(inputXlsxFile, data_only=True)	 
	sheets = wb.get_sheet_names()   # 获取所有表格(worksheet)的名字
	sheet0 = sheets[0]  # 第一个表格的名称
	ws = wb.get_sheet_by_name(sheet0) # 获取特定的 worksheet
	rows = ws.rows
	columns = ws.columns
	# for row in rows:

	# 行迭代
	# print type(rows)

		# print 
	# content = []
	# for row in rows:
	# 	line = [col.value for col in row]
	# 	print line
	# 	content.append(line)
	# pass

	structESize = 0


	columnNameDict ={}
	def initColumnName():
		new_rows = wb.get_sheet_by_name(sheet0).rows
		for row in new_rows:
			columnIndex = 0
			for col in row:
				# if col.value == columnName:
				# 	return columnIndex
				if type(col.value) !=  types.NoneType:
					columnNameDict[col.value.strip()] = columnIndex
				columnIndex = columnIndex+1
			break
		return -1




	def getColumnNameIndex(columnName):
		# print (columnNameDict)
		if columnNameDict.has_key(columnName):
			return columnNameDict[columnName]
		return -1

	def getColumn(columnName):
		if columnNameDict.has_key(columnName):
			return columnNameDict[columnName]
		return None


	initColumnName()
	# print columnNameDict

	rowIndex = 0
	for row in rows:
		# print("rowIndex", rowIndex)
		bRowEffect = True
		for col in row:
			if type(col.value) == types.NoneType:
				bRowEffect = False
			break
		
		#check sort key

		columnIndex =  getColumnNameIndex(sortkeyDict[meta]["cname"])
		# print ("columnIndex", sortkeyDict[meta],columnIndex)
		# print("sortkeyDict[", sortkeyDict[meta]["cname] )
		keyBoolen  = (sortkeyDict[meta]["type"] == "int" and type(row[columnIndex].value) != types.UnicodeType) or (sortkeyDict[meta]["type"] != "int")

		if bRowEffect == True and keyBoolen == True:
			if rowIndex >= 1:
				structE =  dom.createElement(meta)
				for obj in structDict[meta]:
					columnIndex =  getColumnNameIndex(obj["cname"].strip())
					columnE = dom.createElement(obj["name"])

					# print(obj["cname"].strip().encode())
					# print ( obj["cname"])
					# print columnIndex
					# print ( obj["cname"], columnIndex)
					# print (type(row[columnIndex].value) )
					if type(row[columnIndex].value) == types.IntType:
						# print "int--"
						columnE.appendChild(dom.createTextNode("%d"%(row[columnIndex].value)))
					elif type(row[columnIndex].value) == types.LongType:
						# print "long--"
						columnE.appendChild(dom.createTextNode("%d"%(row[columnIndex].value)))
					elif type(row[columnIndex].value) == types.UnicodeType:

						if obj.has_key("size"):
							utf8string = (row[columnIndex].value).encode("utf-8")
							# print ("checksize", obj["size"], len(utf8string))
							if obj["size"] <= len(utf8string):
								raise Exception(u"无效的行!行:%d   列:%s"%(rowIndex,(obj["cname"] )))

						# print("obj value", row[columnIndex].value.strip())
						if obj["type"] == "int":
							columnE.appendChild(dom.createTextNode(row[columnIndex].value.strip("[]")))
						else:
							columnE.appendChild(dom.createTextNode(row[columnIndex].value.strip()))
					elif type(row[columnIndex].value) == types.NoneType:

						if obj["type"] == "int":
							columnE.appendChild(dom.createTextNode("0"))						
						elif obj["type"] == "string":
							columnE.appendChild(dom.createTextNode(""))						
						else:
							columnE.appendChild(dom.createTextNode(" "))						
					else:
						if obj["type"] == "int":
							columnE.appendChild(dom.createTextNode("%d"%(row[columnIndex].value)))
						else:
							columnE.appendChild(dom.createTextNode("%r"%(row[columnIndex].value)))

					structE.appendChild(columnE)
					# print obj["cname"], row[columnIndex].value
				structESize = structESize+1
				root.appendChild(structE)
		rowIndex = rowIndex+1
	rowIndex = 0
	for row in rows:
		if rowIndex == 1:
		# for col in row:
		# 	print col.value
			for obj in structDict[meta]:
				# print getColumnNameIndex(rows, obj["cname"])
				# print obj
				break
		rowIndex = rowIndex +1

	countE = dom.createElement('Count')
	countE.appendChild(dom.createTextNode("%d"%(structESize)))
	resheadE.appendChild(countE)


	# content = []
	# for row in rows:
	#     line = [col.value for col in row]
	#     print line
	#     content.append(line)
	dom.writexml(f, addindent='  ', newl='\n',encoding='utf-8')

def parseXml():
	# print("parseXml", "./convlist")
	# convertListName = "./convlist.xml"

	fileName = "./convlist.xml"
	file_xml = open(fileName,"r").read()
	charset = get_charset(file_xml)
	print "old charset:"+charset
	file_xml = file_xml.decode(charset).encode("utf-8")

	try: 
		ConvListRoot = ET.fromstring(file_xml)     #打开xml文档 
	except Exception, e: 
		print e
		print "Error:cannot parse file:." + fileName
		return

	for child in ConvListRoot:
		print("tag GameFile root", child.tag)

		if child.tag == "ConvTree":
			print("tag convtree")
			for subChild in child:
				if subChild.tag == "CommNode":
					for resNode in subChild:
						# print (resNode.attrib)["BinStyles"]
						# print (resNode.attrib)["Name"]
						# print (resNode.attrib)["Sort"]
						# print (resNode.attrib)["Meta"]
						# print (resNode.attrib)["ExcelFile"]
						# print (resNode.attrib)["BinFile"]
						xlsToXml((resNode.attrib)["ExcelFile"], (resNode.attrib)["Meta"])

if __name__ == '__main__':
	# init_resource_root()

	parseResouceXml()
	parseXml()

	# read_excel()
